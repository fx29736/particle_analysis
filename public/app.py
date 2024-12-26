import streamlit as st
import numpy as np
import pandas as pd
from PIL import Image
Image.MAX_IMAGE_PIXELS = None #読み込みピクセル数の上限を解除
import cv2
import matplotlib.pyplot as plt
import os
import io
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows
import tempfile
st.set_page_config(layout="wide")

def pil2cv(image):
    ''' PIL型 -> OpenCV型 に変換する関数'''
    new_image = np.array(image, dtype=np.uint8)  # PIL画像をNumPy配列に変換
    if new_image.ndim == 2:  # モノクロ画像の場合
        pass
    elif new_image.shape[2] == 3:  # カラー画像の場合 (RGB)
        new_image = cv2.cvtColor(new_image, cv2.COLOR_RGB2BGR)
    elif new_image.shape[2] == 4:  # 透過画像の場合 (RGBA)
        new_image = cv2.cvtColor(new_image, cv2.COLOR_RGBA2BGRA)
    return new_image

def display_image_with_matplotlib(image, title, cmap=None):
    fig, ax = plt.subplots()
    ax.imshow(image, cmap=cmap)
    ax.set_title(title)
    ax.axis('off')
    return fig

def display_histogram(data, title, xlabel, ylabel, bins=20):
    fig, ax = plt.subplots()
    ax.hist(data, bins=bins, color='blue', edgecolor='black')
    ax.set_title(title)
    ax.set_xlabel(xlabel)
    ax.set_ylabel(ylabel)
    return fig

def analyze_image(image_path, x, y, w, h, th, radio):
    img = Image.open(image_path)  # 画像を開く
    cv_img = pil2cv(img)  # OpenCV形式に変換

    # トリミング領域を表示
    trimmed_img_array = cv_img.copy()
    
    # 矩形を表示
    cv2.rectangle(trimmed_img_array, (x, y), (x + w, y + h), (0, 0, 255), 10)
    
    # トリミングエリアの塗りつぶし
    overlay = trimmed_img_array.copy()
    alpha = 0.8  # 透明度 80%
    cv2.rectangle(overlay, (x, y), (x + w, y + h), (0, 0, 255), -1)  # 塗りつぶし
    cv2.addWeighted(overlay, alpha, trimmed_img_array, alpha, 0, trimmed_img_array)

    # トリミング
    x_end = min(x + w, cv_img.shape[1])
    y_end = min(y + h, cv_img.shape[0])
    img_trimmed = cv_img[y:y_end, x:x_end]

    gray = img_trimmed
    if len(img_trimmed.shape) == 3 and img_trimmed.shape[2] != 1:  # カラー画像の場合
        gray = cv2.cvtColor(img_trimmed, cv2.COLOR_BGR2GRAY)

    # 各種バイナリ化処理
    ret, th1 = cv2.threshold(gray, th, 255, cv2.THRESH_BINARY_INV)  # 閾値処理 (黒を粒子として認識)
    th2 = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_MEAN_C,
                                cv2.THRESH_BINARY_INV, 11, 2)  # 適応的閾値処理 (平均法)
    th3 = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                cv2.THRESH_BINARY_INV, 11, 2)  # 適応的閾値処理 (ガウシアン法)

    # 大津の二値化処理
    ret2, th4 = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)

    # ガウシアンフィルタを適用した後に大津の二値化処理
    blur = cv2.GaussianBlur(gray, (5, 5), 0)
    ret3, th5 = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)

    # 選択されたバイナリ処理方法に応じて、処理後の画像を表示
    if radio == "Threshold":
        binary_image = th1
    elif radio == "Adaptive threshold mean":
        binary_image = th2
    elif radio == "Adaptive threshold Gaussian":
        binary_image = th3
    elif radio == "Otsu' thresholding":
        binary_image = th4
    elif radio == "Otsu's thresholding + Gaussian filter":
        binary_image = th5

    num_labels, labels, stats, centroids = cv2.connectedComponentsWithStats(binary_image, connectivity=8, ltype=cv2.CV_32S)
    labeled_image = cv2.cvtColor(cv2.bitwise_not(binary_image), cv2.COLOR_GRAY2BGR)
    
    particle_info = []
    particle_areas = []
    centroid_x_coords = []
    for i in range(1, num_labels):  # 0番目のラベルは背景なので無視
        x, y, width, height, area = stats[i]
        cx, cy = centroids[i]
        offset = 20  # 矩形のサイズを20ピクセルずつ拡張するためのオフセット
        cv2.rectangle(labeled_image, (x - offset, y - offset), (x + width + offset, y + height + offset), (0, 0, 255), 4)  # 矩形の線の太さを4に変更
        cv2.putText(labeled_image, str(i), (x - offset, y - offset - 10), cv2.FONT_HERSHEY_SIMPLEX, 1.5, (0, 0, 255), 3)  # フォントサイズを1.5、太さを3に変更

        # サイズや形状を計算
        contour_points = np.where(labels == i)
        contours = np.array(list(zip(contour_points[1], contour_points[0])))

        if len(contours) >= 5:
            ellipse = cv2.fitEllipse(contours)
            (_, _), (major_axis, minor_axis), _ = ellipse
        else:
            major_axis, minor_axis = 0, 0

        perimeter = cv2.arcLength(contours, True)

        particle_info.append({
            "Label": i,
            "Centroid_X": cx,
            "Centroid_Y": cy,
            "Area": area,
            "Width": width,
            "Height": height,
            "Major Axis": major_axis,
            "Minor Axis": minor_axis,
            "Perimeter": perimeter
        })
        particle_areas.append(area)
        centroid_x_coords.append(cx)

    return trimmed_img_array, labeled_image, pd.DataFrame(particle_info), particle_areas, centroid_x_coords

# サイドバーにフォルダパスの入力UIと他の入力UIを設置
with st.sidebar:
    folder_path = st.text_input('フォルダパスを入力してください', '')
    st.divider()
    st.text("解析範囲を下記で指定してください。範囲は解析前画像に矩形で表示されます")
    x = st.number_input('Trimming start position x', min_value=0, max_value=None, value=100)
    y = st.number_input('Trimming start position y', min_value=0, max_value=None, value=100)
    w = st.number_input('Trimming width', min_value=1, max_value=None, value=200)
    h = st.number_input('Trimming height', min_value=1, max_value=None, value=200)
    st.divider()
    radio = st.radio(
        "Choose a binary method",
        ("Threshold", "Adaptive threshold mean", "Adaptive threshold Gaussian",
         "Otsu' thresholding", "Otsu's thresholding + Gaussian filter")
    )
    th = st.slider('Threshold value', 0, 255, 125)  # 閾値スライダー
    st.divider()
    st.text("ボタンを押すと解析結果のExcelが画像と同じフォルダに保存されます")
    save_button = st.button('解析結果を保存')

# タイトルと内容の説明
st.title('粒子解析')
st.write("指定フォルダ内の全ての画像について解析を行います。トリミング⇒二値化⇒粒子解析")

all_results = {}

# フォルダ内の画像を取得
if folder_path:
    images = [f for f in os.listdir(folder_path) if f.endswith(('.png', '.jpg', '.jpeg'))]
    if images:
        tab_names = images
        tabs = st.tabs(tab_names)

        for tab, image_name in zip(tabs, images):
            with tab:
                st.header(f"Analysis for {image_name}")
                image_path = os.path.join(folder_path, image_name)

                trimmed_img_array, labeled_image, particle_info_df, particle_areas, centroid_x_coords = analyze_image(
                    image_path, x, y, w, h, th, radio
                )

                st.image(cv2.cvtColor(trimmed_img_array, cv2.COLOR_BGR2RGB), caption='元画像 (トリミング領域表示)', use_container_width=True)
                labeled_image_fig = display_image_with_matplotlib(labeled_image, 'Labeled Particles')
                particle_sizes_fig = display_histogram(particle_areas, 'Histogram of Particle Sizes', 'Particle Area', 'Frequency')
                x_coords_fig = display_histogram(centroid_x_coords, 'Histogram of Particle X Coordinates', 'Centroid X Coordinate', 'Frequency')
                
                # 保存用の結果を集める
                all_results[image_name] = {
                    'labeled_image': labeled_image_fig,
                    'particle_sizes': particle_sizes_fig,
                    'x_coords': x_coords_fig,
                    'particle_info_df': particle_info_df,
                    'x_coords_data': centroid_x_coords
                }

                st.pyplot(labeled_image_fig)
                st.pyplot(particle_sizes_fig)
                st.pyplot(x_coords_fig)
                st.write('## Particle Information')
                st.write(particle_info_df)

# 解析結果の保存ボタンを押したときの処理
if save_button and all_results:
    st.sidebar.write("解析結果の保存を開始します...")
    
    wb = Workbook()
    wb.remove(wb.active)  # デフォルトのシートを削除
    
    for image_name, results in all_results.items():
        ws = wb.create_sheet(title=image_name)

        # labeled_image
        labeled_image_buf = io.BytesIO()
        results['labeled_image'].savefig(labeled_image_buf, format='png')
        labeled_image_buf.seek(0)
        img = XLImage(labeled_image_buf)
        img.anchor = 'A1'
        ws.add_image(img)

        # Histogram of Particle Sizes
        particle_sizes_buf = io.BytesIO()
        results['particle_sizes'].savefig(particle_sizes_buf, format='png')
        particle_sizes_buf.seek(0)
        img = XLImage(particle_sizes_buf)
        img.anchor = 'K1'
        ws.add_image(img)

        # Histogram of Particle X Coordinates
        x_coords_buf = io.BytesIO()
        results['x_coords'].savefig(x_coords_buf, format='png')
        x_coords_buf.seek(0)
        img = XLImage(x_coords_buf)
        img.anchor = 'U1'
        ws.add_image(img)

        # Particle Information table
        for r_idx, row in enumerate(dataframe_to_rows(results['particle_info_df'], index=False, header=True), 30):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

    # ファイル保存のチェック
    try:
        # 保存先ディレクトリが存在しない場合は作成
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
            st.sidebar.write(f"{folder_path} を作成しました。")
        
        # ファイルパスを指定
        file_name = "analysis_results.xlsx"
        file_path = os.path.join(folder_path, file_name)
        
        # Excelファイルを保存
        wb.save(file_path)
        st.sidebar.write(f"ファイルを保存しました: {file_path}")
    except Exception as e:
        st.sidebar.write(f"ファイルの保存に失敗しました: {e}")
